import socket
import threading
import openpyxl
import datetime
import hashlib  # Importing hashlib for SHA-256
import pickle

# Load workbooks
loginInfo_Workbook = openpyxl.load_workbook('database.xlsx')
userData = loginInfo_Workbook.active

chat_Workbook = openpyxl.load_workbook('chatHistory.xlsx')
historyData = chat_Workbook.active

# Function for checking validity with hashed password using SHA-256
def login(username, password, client):
    for i in range(2, userData.max_row + 1):
        if userData.cell(i, 1).value == username:
            stored_hashed_password = userData.cell(i, 3).value
            # Hash the entered password and compare it with stored hash
            hashed_password = hashlib.sha256(password.encode('utf-8')).hexdigest()
            if hashed_password == stored_hashed_password:
                client.send("Login Successful".encode())
                return True
            else:
                return False
    return False

# Function for registration with SHA-256 hashed password
def signup(username, password, client, email):
    i = userData.max_row + 1
    # Hash the password with SHA-256
    hashed_password = hashlib.sha256(password.encode('utf-8')).hexdigest()
    userData.cell(i, 1).value = username
    userData.cell(i, 2).value = email
    userData.cell(i, 3).value = hashed_password
    client.send("success".encode())
    loginInfo_Workbook.save("database.xlsx")
    return True

# Connection Data
host = ''   # host IP address
port = 5555

# Starting Server
server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server.bind((host, port))
server.listen()

# Lists For Clients and Their Nicknames
clients = []
nicknames = []

# Sending Messages To All Connected Clients
def broadcast(message):
    message = message.decode()
    current_time = datetime.datetime.now()
    string_time = current_time.strftime("%H:%M")
    message = "\n" + string_time + "\n" + message
    message = message.encode()
    for client in clients:
        client.send(message)

# Handling Messages From Clients
def handle(client):
    while True:
        try:
            # Broadcasting Messages
            message = client.recv(1024)

            if message == 'GIVE'.encode('ascii'):
                for user in nicknames:
                    client.send(user.encode('ascii'))
            
            else:   
                broadcast(message)
                i = historyData.max_row + 1
                historyData.cell(i, 1).value = datetime.datetime.now()
                historyData.cell(i, 2).value = message.decode().split(': ')[0]
                historyData.cell(i, 3).value = message.decode().split(': ')[1]
        
        except:
            # Removing And Closing Clients
            index = clients.index(client)
            clients.remove(client)
            client.close()
            nickname = nicknames[index]
            broadcast('{} left!'.format(nickname).encode('ascii'))
            nicknames.remove(nickname)
            break

# Receiving / Listening Function
def receive():
    while True:
        # Accept Connection
        client, address = server.accept()
        print("Connected with {}".format(str(address)))

        # Request And Store Nickname
        mode = client.recv(1024).decode('ascii')  # listening to client whether they want to sign in or sign up

        if mode == "Login":
            client.send('NICK'.encode('ascii'))
            nickname = client.recv(1024).decode('ascii')
            client.send('PASS'.encode('ascii'))
            password = client.recv(1024).decode('ascii')
            flag = login(nickname, password, client)
        
        elif mode == "signup":
            client.send('NICK'.encode('ascii'))
            nickname = client.recv(1024).decode('ascii')
            client.send('PASS'.encode('ascii'))
            password = client.recv(1024).decode('ascii')
            client.send('EMAIL'.encode('ascii'))
            email = client.recv(1024).decode('ascii')
            flag = signup(nickname, password, client, email)

        if flag:
            # Adding the username in the list
            nicknames.append('User:' + nickname)
            clients.append(client)

            # Showing connection message
            print("Connected with {}".format(str(address)))
            print("Nickname is {}".format(nickname))

            # Sending confirmation message to the client
            client.send('Login Success'.encode('ascii'))

            # Sending the previous history to the newly joined client
            for i in range(2, historyData.max_row + 1):
                time = historyData.cell(i, 1).value.strftime("%H:%M")
                client.send(f"{time} \n{historyData.cell(i, 2).value} : {historyData.cell(i, 3).value}\n\n".encode())

            broadcast("{} joined!\n".format(nickname).encode('ascii'))

            # Start Handling Thread For Client
            thread = threading.Thread(target=handle, args=(client,))
            thread.start()
        
        else:
            client.send('Wrong INFO'.encode('ascii'))

print("[STARTING SERVER]: Server is listening ...")
receive()
